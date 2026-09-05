#!/usr/bin/env python3
"""
make_workbook.py -- Build the Excel comparison workbook from benchmark.csv.

Everything derived (speed-up, efficiency, overhead ratios) is written as a
*formula* referencing the raw-data sheet, never as a value computed here. Edit
a measured time on the "Raw Data" sheet and the whole workbook recalculates,
which is what makes it auditable rather than just a picture of our numbers.

Usage: python3 bench/make_workbook.py [results/benchmark.csv] [out.xlsx]
"""
import csv
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

CSV_PATH = sys.argv[1] if len(sys.argv) > 1 else "results/benchmark.csv"
OUT      = sys.argv[2] if len(sys.argv) > 2 else "results/Comparison_Sequential_vs_MPI.xlsx"

# ---- house style -------------------------------------------------------
FONT      = "Arial"
INK       = "1F2328"
ACCENT    = "0F6E77"
BLUE      = "0000FF"          # hardcoded input (a measured time)
GOOD      = "1F7A46"
BAD       = "B03A2E"
HDR_FILL  = PatternFill("solid", fgColor="0F6E77")
SUB_FILL  = PatternFill("solid", fgColor="E4ECEC")
WARN_FILL = PatternFill("solid", fgColor="FDECEA")
GOOD_FILL = PatternFill("solid", fgColor="E7F4EC")
THIN      = Side(style="thin", color="C9CED4")
BOX       = Border(bottom=THIN)

F_MS   = '0.000'
F_X    = '0.00"x"'
F_PCT  = '0.0%'
F_INT  = '#,##0'

def title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=14, bold=True, color=INK)

def note(ws, cell, text, italic=True):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=9, italic=italic, color="6A737D")

def header_row(ws, row, labels, start=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=lab)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def main():
    with open(CSV_PATH) as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()

    # =====================================================================
    # 1. README
    # =====================================================================
    ws = wb.active
    ws.title = "README"
    widths(ws, {"A": 34, "B": 96})
    title(ws, "A1", "Distributed Image Greyscale — Sequential vs. MPI")
    note(ws, "A2", "Comparison workbook. Every derived figure is a live formula over the Raw Data sheet.")

    info = [
        ("", ""),
        ("SHEETS", ""),
        ("Raw Data", "Every measured run, one row per configuration. Blue = measured input. Source: results/benchmark.csv"),
        ("Compute Comparison", "Compute-phase time, speed-up and parallel efficiency vs. the sequential baseline."),
        ("End-to-End Comparison", "Whole-program time including read, communication and write — what a user actually waits for."),
        ("Overhead Analysis", "Communication and I/O cost set against the computation they were meant to accelerate."),
        ("MPI Startup Cost", "Fixed cost of MPI_Init + MPI_Finalize, excluded from all timings above."),
        ("", ""),
        ("HOW TO READ IT", ""),
        ("Blue figures", "Measured inputs. Change one and the workbook recalculates."),
        ("Black figures", "Formulas derived from the blue inputs."),
        ("Green / red fills", "Faster than sequential / slower than sequential."),
        ("Compute speed-up", "T_sequential(compute) / T_parallel(compute) — judges the decomposition."),
        ("End-to-end speed-up", "T_sequential(total) / T_parallel(total) — judges the program. The honest number."),
        ("Efficiency", "Compute speed-up divided by process count. 100% = perfect scaling."),
        ("", ""),
        ("METHOD", ""),
        ("Timing", "Kernel repeated 60x in-process; best of 9 whole-program trials. The minimum is used, not the mean."),
        ("Parallel phases", "Reduced across ranks with MPI_MAX: a phase costs what its slowest rank costs."),
        ("Repetition loop", "Subtracted from the total, so sequential and MPI totals are directly comparable."),
        ("Correctness", "Every run re-verified byte-for-byte against the sequential output. See logs/02_correctness.log."),
        ("", ""),
        ("EVIDENCE", ""),
        ("Raw program output", "03_Logs_and_Data/logs/ — one untouched log per configuration."),
        ("Written analysis", "REPORT.md, sections 5 and 6."),
        ("Source data", "03_Logs_and_Data/benchmark.csv, environment.txt"),
    ]
    r = 3
    for k, v in info:
        if k and not v:
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
        elif k:
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10, bold=True, color=INK)
            c = ws.cell(row=r, column=2, value=v)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(vertical="top")
        r += 1

    # =====================================================================
    # 2. Raw Data  (the single source of truth)
    # =====================================================================
    rd = wb.create_sheet("Raw Data")
    widths(rd, {"A": 15, "B": 8, "C": 9, "D": 9, "E": 13, "F": 11, "G": 11,
                "H": 11, "I": 11, "J": 11, "K": 11})
    title(rd, "A1", "Raw measurements")
    note(rd, "A2", "All times in seconds, best of 9 trials. Measured inputs shown in blue — every other "
                   "sheet is formulas over this one. Source: results/benchmark.csv")
    cols = ["mode", "procs", "width", "height", "pixels", "read_s", "scatter_s",
            "compute_s", "gather_s", "write_s", "total_s"]
    header_row(rd, 4, cols)

    # csv row index -> worksheet row, so other sheets can reference precisely
    ref = {}
    r = 5
    for row in rows:
        key = (row["mode"], int(row["procs"]), f'{row["width"]}x{row["height"]}')
        ref[key] = r
        for i, c in enumerate(cols, start=1):
            v = row[c]
            if c == "mode":
                cell = rd.cell(row=r, column=i, value=v)
                cell.font = Font(name=FONT, size=10, color=INK)
            elif c in ("procs", "width", "height", "pixels"):
                cell = rd.cell(row=r, column=i, value=int(v))
                cell.font = Font(name=FONT, size=10, color=INK)
                cell.number_format = F_INT
            else:
                cell = rd.cell(row=r, column=i, value=float(v) if v else None)
                cell.font = Font(name=FONT, size=10, color=BLUE)
                cell.number_format = '0.000000'
            cell.border = BOX
        r += 1
    last_data_row = r - 1
    note(rd, f"A{r + 1}", "Empty scatter/gather cells: that phase does not exist in that mode "
                          "(the sequential program and the MPI-IO path never move data between ranks).")

    resolutions = []
    for row in rows:
        res = f'{row["width"]}x{row["height"]}'
        if res not in resolutions:
            resolutions.append(res)
    procs = sorted({int(x["procs"]) for x in rows if x["mode"] != "sequential"})
    MODES = [("mpi-scatter", "MPI scatter/gather"), ("mpi-mpiio", "MPI collective I/O")]

    def cell(mode, np_, res, col):
        """A1-style reference into Raw Data for one measured value."""
        return f"'Raw Data'!{get_column_letter(col)}{ref[(mode, np_, res)]}"

    C_READ, C_SCAT, C_COMP, C_GATH, C_WRITE, C_TOT = 6, 7, 8, 9, 10, 11

    # =====================================================================
    # 3. Compute Comparison
    # =====================================================================
    cs = wb.create_sheet("Compute Comparison")
    widths(cs, {"A": 24, "B": 8, "C": 14, "D": 14, "E": 14, "F": 13, "G": 30})
    title(cs, "A1", "Compute phase — how well the decomposition scales")
    note(cs, "A2", "Speed-up = sequential compute time / parallel compute time. Efficiency = speed-up / ranks. "
                   "This measures the parallel work only; see 'End-to-End Comparison' for the whole program.")
    row = 4
    chart_anchors = []
    for res in resolutions:
        px = int(next(x for x in rows if f'{x["width"]}x{x["height"]}' == res)["pixels"])
        c = cs.cell(row=row, column=1, value=f"{res}   ({px/1e6:.1f} MPixel)")
        c.font = Font(name=FONT, size=11, bold=True, color=ACCENT)
        cs.cell(row=row, column=1).fill = SUB_FILL
        row += 1
        header_row(cs, row, ["Mode", "Ranks", "Compute (ms)", "Speed-up",
                             "Efficiency", "Ideal", "Comment"])
        hdr = row
        row += 1
        seq = cell("sequential", 1, res, C_COMP)
        # baseline row
        cs.cell(row=row, column=1, value="Sequential baseline").font = Font(name=FONT, size=10, bold=True)
        cs.cell(row=row, column=2, value=1).font = Font(name=FONT, size=10)
        cs.cell(row=row, column=3, value=f"={seq}*1000").number_format = F_MS
        cs.cell(row=row, column=4, value=1.0).number_format = F_X
        cs.cell(row=row, column=5, value=1.0).number_format = F_PCT
        cs.cell(row=row, column=6, value=1.0).number_format = F_X
        for col in range(1, 8):
            cs.cell(row=row, column=col).font = Font(
                name=FONT, size=10, bold=(col == 1), color=INK)
            cs.cell(row=row, column=col).border = BOX
        row += 1
        first_series = row
        for mode, label in MODES:
            for np_ in procs:
                cs.cell(row=row, column=1, value=label)
                cs.cell(row=row, column=2, value=np_)
                cs.cell(row=row, column=3, value=f"={cell(mode, np_, res, C_COMP)}*1000")
                cs.cell(row=row, column=4, value=f"=$C${row-0}").value = f"={seq}*1000/C{row}"
                cs.cell(row=row, column=5, value=f"=D{row}/B{row}")
                cs.cell(row=row, column=6, value=f"=B{row}")
                cs.cell(row=row, column=3).number_format = F_MS
                cs.cell(row=row, column=4).number_format = F_X
                cs.cell(row=row, column=5).number_format = F_PCT
                cs.cell(row=row, column=6).number_format = F_X
                # 1.02 rather than 1.00: a couple of tenths of a percent is measurement
                # noise, not a cache effect, and labelling it as one would overclaim.
                cs.cell(row=row, column=7, value=f'=IF(E{row}>=1.02,"superlinear — cache effect",'
                                                 f'IF(E{row}>=0.95,"near-ideal scaling",'
                                                 f'IF(E{row}>=0.9,"good scaling","efficiency loss")))')
                for col in range(1, 8):
                    cs.cell(row=row, column=col).font = Font(name=FONT, size=10, color=INK)
                    cs.cell(row=row, column=col).border = BOX
                cs.cell(row=row, column=4).font = Font(name=FONT, size=10, bold=True, color=GOOD)
                row += 1
        chart_anchors.append((res, hdr, first_series, row - 1))
        row += 2

    # one chart per resolution: measured speed-up vs. ideal, scatter/gather series
    for res, hdr, first, last in chart_anchors:
        ch = LineChart()
        ch.title = f"Compute speed-up — {res}"
        ch.style = 2
        ch.y_axis.title = "Speed-up"
        ch.x_axis.title = "MPI ranks"
        ch.height, ch.width = 7.2, 13
        n = len(procs)
        data = Reference(cs, min_col=4, max_col=4, min_row=first, max_row=first + n - 1)
        ideal = Reference(cs, min_col=6, max_col=6, min_row=first, max_row=first + n - 1)
        cats = Reference(cs, min_col=2, max_col=2, min_row=first, max_row=first + n - 1)
        ch.add_data(data, titles_from_data=False)
        ch.add_data(ideal, titles_from_data=False)
        ch.set_categories(cats)
        ch.series[0].tx = None
        cs.add_chart(ch, f"I{hdr}")

    # =====================================================================
    # 4. End-to-End Comparison
    # =====================================================================
    es = wb.create_sheet("End-to-End Comparison")
    widths(es, {"A": 24, "B": 8, "C": 13, "D": 13, "E": 13, "F": 13, "G": 14, "H": 26})
    title(es, "A1", "Whole program — what the user actually waits for")
    note(es, "A2", "Read + communication + compute + write. A speed-up below 1.00x means the parallel "
                   "program is SLOWER than the sequential one. Green = faster, red = slower.")
    row = 4
    for res in resolutions:
        c = es.cell(row=row, column=1, value=res)
        c.font = Font(name=FONT, size=11, bold=True, color=ACCENT)
        c.fill = SUB_FILL
        row += 1
        header_row(es, row, ["Mode", "Ranks", "I/O (ms)", "Comm (ms)", "Compute (ms)",
                             "Total (ms)", "Speed-up", "Verdict"])
        row += 1
        seq_tot = cell("sequential", 1, res, C_TOT)
        es.cell(row=row, column=1, value="Sequential baseline").font = Font(name=FONT, size=10, bold=True)
        es.cell(row=row, column=2, value=1)
        es.cell(row=row, column=3, value=f"=({cell('sequential',1,res,C_READ)}+{cell('sequential',1,res,C_WRITE)})*1000")
        es.cell(row=row, column=4, value=0)
        es.cell(row=row, column=5, value=f"={cell('sequential',1,res,C_COMP)}*1000")
        es.cell(row=row, column=6, value=f"={seq_tot}*1000")
        es.cell(row=row, column=7, value=1.0)
        es.cell(row=row, column=8, value="baseline")
        for col in range(3, 7):
            es.cell(row=row, column=col).number_format = F_MS
        es.cell(row=row, column=7).number_format = F_X
        for col in range(1, 9):
            es.cell(row=row, column=col).font = Font(name=FONT, size=10, bold=(col == 1), color=INK)
            es.cell(row=row, column=col).border = BOX
        row += 1
        for mode, label in MODES:
            for np_ in procs:
                es.cell(row=row, column=1, value=label)
                es.cell(row=row, column=2, value=np_)
                es.cell(row=row, column=3, value=f"=({cell(mode,np_,res,C_READ)}+{cell(mode,np_,res,C_WRITE)})*1000")
                es.cell(row=row, column=4, value=f"=(N({cell(mode,np_,res,C_SCAT)})+N({cell(mode,np_,res,C_GATH)}))*1000")
                es.cell(row=row, column=5, value=f"={cell(mode,np_,res,C_COMP)}*1000")
                es.cell(row=row, column=6, value=f"={cell(mode,np_,res,C_TOT)}*1000")
                es.cell(row=row, column=7, value=f"={seq_tot}*1000/F{row}")
                es.cell(row=row, column=8, value=f'=IF(G{row}>=1,"faster than sequential","SLOWER than sequential")')
                for col in range(3, 7):
                    es.cell(row=row, column=col).number_format = F_MS
                es.cell(row=row, column=7).number_format = F_X
                for col in range(1, 9):
                    cc = es.cell(row=row, column=col)
                    cc.font = Font(name=FONT, size=10, color=INK)
                    cc.border = BOX
                row += 1
        row += 2

    # =====================================================================
    # 5. Overhead Analysis
    # =====================================================================
    ov = wb.create_sheet("Overhead Analysis")
    widths(ov, {"A": 16, "B": 22, "C": 8, "D": 17, "E": 17, "F": 17, "G": 19, "H": 26})
    title(ov, "A1", "Does the data movement cost more than the work it distributes?")
    note(ov, "A2", "The kernel is memory-bandwidth bound (1.5 ops per byte). If moving the image costs more "
                   "than converting it, no amount of compute scaling can win. Ratio > 1.00 means the "
                   "communication alone exceeds the entire sequential computation.")
    header_row(ov, 4, ["Resolution", "Mode", "Ranks", "Seq. compute (ms)",
                       "Communication (ms)", "Parallel compute (ms)",
                       "Comm / seq compute", "Conclusion"])
    row = 5
    for res in resolutions:
        for mode, label in MODES:
            for np_ in procs:
                seq = cell("sequential", 1, res, C_COMP)
                ov.cell(row=row, column=1, value=res)
                ov.cell(row=row, column=2, value=label)
                ov.cell(row=row, column=3, value=np_)
                ov.cell(row=row, column=4, value=f"={seq}*1000")
                ov.cell(row=row, column=5, value=f"=(N({cell(mode,np_,res,C_SCAT)})+N({cell(mode,np_,res,C_GATH)}))*1000")
                ov.cell(row=row, column=6, value=f"={cell(mode,np_,res,C_COMP)}*1000")
                ov.cell(row=row, column=7, value=f"=IF(E{row}=0,0,E{row}/D{row})")
                ov.cell(row=row, column=8, value=f'=IF(E{row}=0,"no data movement at all",'
                                                f'IF(G{row}>1,"communication exceeds the work","communication is affordable"))')
                for col in (4, 5, 6):
                    ov.cell(row=row, column=col).number_format = F_MS
                ov.cell(row=row, column=7).number_format = F_X
                for col in range(1, 9):
                    cc = ov.cell(row=row, column=col)
                    cc.font = Font(name=FONT, size=10, color=INK)
                    cc.border = BOX
                row += 1
    note(ov, f"A{row+1}", "Rows with zero communication are the MPI-IO path: every rank reads and writes only "
                          "its own byte range, so no pixel ever crosses a rank boundary.")

    # =====================================================================
    # 6. MPI Startup Cost
    # =====================================================================
    st = wb.create_sheet("MPI Startup Cost")
    widths(st, {"A": 10, "B": 22, "C": 22, "D": 24, "E": 30})
    title(st, "A1", "The bill before the first pixel")
    note(st, "A2", "MPI_Init + MPI_Finalize, measured by tools/mpi_startup.c, best of 3. This cost is EXCLUDED "
                   "from every timing in this workbook — the timed region starts after MPI_Init — so the "
                   "speed-ups elsewhere are generous to MPI. Source: logs/01_environment.log")
    header_row(st, 4, ["Ranks", "Startup (ms)", "HD conversion (ms)",
                       "Startup / conversion", "Meaning"])
    startup = {1: 291.995, 2: 327.197, 3: 364.955, 4: 391.592}
    hd_seq_comp = cell("sequential", 1, resolutions[0], C_COMP)
    row = 5
    for np_ in sorted(startup):
        st.cell(row=row, column=1, value=np_)
        c = st.cell(row=row, column=2, value=startup[np_])
        c.font = Font(name=FONT, size=10, color=BLUE)
        c.number_format = F_MS
        st.cell(row=row, column=3, value=f"={hd_seq_comp}*1000").number_format = F_MS
        st.cell(row=row, column=4, value=f"=B{row}/C{row}").number_format = F_X
        st.cell(row=row, column=5, value=f'=TEXT(D{row},"0")&"x the entire HD conversion"')
        for col in range(1, 6):
            cc = st.cell(row=row, column=col)
            if col != 2:
                cc.font = Font(name=FONT, size=10, color=INK)
            cc.border = BOX
        row += 1
    note(st, f"A{row+1}", "Startup figures are measured inputs (blue). Column C and D are formulas over the "
                          "Raw Data sheet, so they follow any change to the measured HD compute time.")

    bar = BarChart()
    bar.title = "MPI fixed startup cost vs. the work it is meant to accelerate"
    bar.y_axis.title = "Milliseconds"
    bar.x_axis.title = "MPI ranks"
    bar.height, bar.width = 8, 15
    bar.add_data(Reference(st, min_col=2, max_col=3, min_row=4, max_row=row - 1),
                 titles_from_data=True)
    bar.set_categories(Reference(st, min_col=1, min_row=5, max_row=row - 1))
    st.add_chart(bar, "G4")

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
    wb.active = 0

    os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(rows)} raw rows, {len(wb.worksheets)} sheets)")


if __name__ == "__main__":
    main()
