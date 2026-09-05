#!/usr/bin/env python3
"""
export_csv_tables.py -- Write plain-CSV copies of the workbook's derived tables.

The .xlsx is the primary artefact, but a reader without Excel should still be
able to check any number. These CSVs hold the *computed* values, exported after
the workbook was recalculated, so they cannot drift from it.

Usage: python3 bench/export_csv_tables.py <output_dir> [workbook.xlsx]
"""
import csv
import sys
import os
from openpyxl import load_workbook

OUT = sys.argv[1]
WB  = sys.argv[2] if len(sys.argv) > 2 else "results/Comparison_Sequential_vs_MPI.xlsx"

SHEETS = {
    "Compute Comparison":     "table_compute_comparison.csv",
    "End-to-End Comparison":  "table_end_to_end_comparison.csv",
    "Overhead Analysis":      "table_overhead_analysis.csv",
    "MPI Startup Cost":       "table_mpi_startup_cost.csv",
}


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:.6g}"
    return str(v)


def main():
    wb = load_workbook(WB, data_only=True)
    os.makedirs(OUT, exist_ok=True)
    for sheet, fname in SHEETS.items():
        ws = wb[sheet]
        path = os.path.join(OUT, fname)
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            for row in ws.iter_rows():
                cells = [clean(c.value) for c in row]
                # Drop the fully empty spacer rows the layout uses between blocks.
                if any(c for c in cells):
                    while cells and cells[-1] == "":
                        cells.pop()
                    w.writerow(cells)
        print(f"  {fname}")


if __name__ == "__main__":
    main()
